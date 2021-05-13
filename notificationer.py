from binance.client import Client
import time
import logging
import telegram_send
import requests
from threading import Thread
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters
import pandas as pd
import atexit
import os.path
import os
from openpyxl import load_workbook
from tradingview_ta import TA_Handler, Interval, Compute
import time
from datetime import datetime
import pickle

api_key = "api_key"
api_secret = "api_secret"
chat_id = "-1001298181916"

class crypto_watch():
    def __init__(self, symbol, time_interval):
        self.symbol = symbol
        self.time_interval = time_interval
        self.first_analysis = get_signal(self.symbol, self.time_interval)
    
    def get_current_analysis(self):
        return get_signal(self.symbol, self.time_interval)
    
    def update_first_analysis(self):
        self.first_analysis = get_signal(self.symbol, self.time_interval)

def get_signal(symbol, time_interval):
    ta_analysis_handler = TA_Handler()
    ta_analysis_handler.set_exchange_as_crypto_or_stock("BINANCE")
    ta_analysis_handler.set_screener_as_crypto()
    ta_analysis_handler.set_symbol_as(symbol=symbol)
    ta_analysis_handler.set_interval_as(time_interval)
    analysis = ta_analysis_handler.get_analysis().summary["RECOMMENDATION"]
    return analysis

def get_current_price(tag):
    client = Client(api_key=api_key, api_secret=api_secret)
    current_price = float((client.get_ticker(symbol= tag))["lastPrice"])
    return current_price

#A HELPER FUNCTION THAT I FOUND IN STACKOVERFLOW
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')#pylint: disable=abstract-class-instantiated

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def check_watch_list():
    while True:
        if len(watch_list) > 0:
            try:
                print("Controlling trackers...")
                for item in watch_list.copy():
                    if isinstance(item, crypto_watch):
                        print(f"{item.symbol} {item.time_interval} First analysis : {item.first_analysis} Current analysis: {item.get_current_analysis()}")
                        if item.first_analysis != item.get_current_analysis():
                            item.update_first_analysis()
                            send_signal_notification(item)
                        else:
                            pass
            except:
                raise Exception
        time.sleep(2)

def find_crypto_var(crypto_symbol, crypto_time):
    crypto_var = None
    for item in watch_list:
        if isinstance(item, crypto_watch):
            if item.symbol == crypto_symbol:
                if crypto_time.lower() == item.time_interval:
                    crypto_var = item
    return crypto_var

def remove_from_watch_list(update, context):
    crypto_symbol, crypto_time = context.args[0], context.args[1].lower()
    crypto_var = find_crypto_var(crypto_symbol, crypto_time=crypto_time)
    if isinstance(crypto_var, crypto_watch):
        watch_list.remove(crypto_var)
        msg = f"{crypto_symbol} {crypto_time} succesfully removed from Tracker-List."
        telegram_send.send(messages=[msg])
        base_url = f"https://api.telegram.org/bot1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU/sendMessage?chat_id=-{chat_id}&text={msg}"
        requests.get(base_url)
    elif crypto_var == None:
        msg = f"{crypto_symbol} {crypto_time} cannot be found."
        telegram_send.send(messages=[msg])
        base_url = f"https://api.telegram.org/bot1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU/sendMessage?chat_id=-{chat_id}&text={msg}"
        requests.get(base_url)

def send_signal_notification(item):
    if isinstance(item, crypto_watch):
        msg = f"RING RINGGG {item.get_current_analysis()} {item.symbol} {item.time_interval} signal has arrived."
        telegram_send.send(messages=[msg])
        base_url = f"https://api.telegram.org/bot1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU/sendMessage?chat_id=-{chat_id}&text={msg}"
        requests.get(base_url)

def check_alarms():
    if len(alarm_list) > 0:
        while True:
            print("Controlling the alarms...")
            for item in alarm_list.copy():
                tag, alarm_price, alarm_type = item[0], item[1], item[2]
                alarm_price_number = float(alarm_price)
                current_price = get_current_price(tag)
                print(f"{tag} crosses {current_price}")
                if alarm_type == "DOWN":
                    if current_price <= alarm_price_number:
                        send_notifications(tag, current_price)
                        alarm_element = [tag, alarm_price, alarm_type]
                        alarm_list.remove(alarm_element)
                elif alarm_type == "UP":
                    if current_price >= alarm_price_number:
                        send_notifications(tag, current_price)
                        alarm_element = [tag, alarm_price, alarm_type]
                        alarm_list.remove(alarm_element)
            time.sleep(2)

def remove_alarm(update, context):
    var = None
    if len(context.args) > 0:
        tag, alarm_price, alarm_type = context.args[0], context.args[1], context.args[2]
        alarm_element = [tag, alarm_price, alarm_type]
        if alarm_element in alarm_list:
            alarm_list.remove(alarm_element)
            print(f"{tag}, {alarm_price}, {alarm_type} has been deleted.")
            var = True
            send_removed_notification(tag, alarm_price, alarm_type, var)
        else:
            var = False
            send_removed_notification(tag, alarm_price, alarm_type, var)
    else:
        msg = "Please write correctly. ./remove_alarm 'COIN_TAG' 'ALARM_PRICE' 'ALARM_TYPE'"
        telegram_send.send(messages=[msg])
        #base_url = f"https://api.telegram.org/bot1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU/sendMessage?chat_id=-{chat_id}&text={msg}"
        #requests.get(base_url)

def send_removed_notification(tag, alarm_price, alarm_type, var):
    if var:
        msg = f"{tag} {alarm_price} {alarm_type} has been deleted from Alarm-List."
        telegram_send.send(messages=[msg])
        #base_url = f"https://api.telegram.org/bot1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU/sendMessage?chat_id=-{chat_id}&text={msg}"
        #requests.get(base_url)
    elif var == False:
        msg = f"{tag} {alarm_price} {alarm_type} could not have been deleted. Some errors occured."
        telegram_send.send(messages=[msg])
        #base_url = f"https://api.telegram.org/bot1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU/sendMessage?chat_id=-{chat_id}&text={msg}"
        #requests.get(base_url)

def add_new_alarm(update, context):
    if len(context.args) == 3:
        tag, alarm_price, alarm_type = context.args[0], context.args[1], context.args[2]
        print(f"{tag}, {alarm_price}, {alarm_type}")
        if tag.replace("USDT", "") in crypto_list():
            if alarm_type == "DOWN" or alarm_type == "UP":
                if ((tag, alarm_price, alarm_type)) not in alarm_list:
                    alarm_list.append([tag, alarm_price, alarm_type])
                    send_notifications_reply(tag, alarm_price)
                else:
                    send_error_notification_alarm_list(tag, alarm_price)
            else:
                send_error_notification_alarm_type()
        else:
            send_error_notification(tag, alarm_price)
    else:
        msg = "Please, use ./add_alarm 'COIN_TAG', 'ALARM_PRICE', 'ALARM_TYPE'"
        telegram_send.send(messages=[msg])
        #base_url = f"https://api.telegram.org/bot1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU/sendMessage?chat_id=-{chat_id}&text={msg}"
        #requests.get(base_url)

def review_alarms(update, context):
    if len(alarm_list) != 0:  
        for item in alarm_list:
            tag = item[0]
            alarm_price = item[1]
            alarm_type = item[2]
            msg = f"Alarm will ring when {tag} crosses {alarm_price}. Alarm-type: {alarm_type}"
            telegram_send.send(messages=[msg])
            #base_url = f"https://api.telegram.org/bot1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU/sendMessage?chat_id=-{chat_id}&text={msg}"
            #requests.get(base_url)
    else:
        msg = "Alarm List is empty."
        #base_url = f"https://api.telegram.org/bot1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU/sendMessage?chat_id=-{chat_id}&text={msg}"
        #requests.get(base_url)
def send_notifications(tag, current_price):
    msg = f"RING RING, {tag} crosses {current_price}."
    #telegram part
    
    telegram_send.send(messages=[msg])
    #base_url = f"https://api.telegram.org/bot1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU/sendMessage?chat_id=-{chat_id}&text={msg}"
    #requests.get(base_url)

def send_error_notification_alarm_type():
    msg = 'Wrong alarm type please write as "UP" or "DOWN"'
    telegram_send.send(messages=[msg])
    #base_url = f"https://api.telegram.org/bot1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU/sendMessage?chat_id=-{chat_id}&text={msg}"
    #requests.get(base_url)

def send_error_notification_alarm_list(tag, alarm_price):
    msg = f"You have already set the alarm on {tag} at {alarm_price}"
    #base_url = f"https://api.telegram.org/bot1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU/sendMessage?chat_id=-{chat_id}&text={msg}"
    #requests.get(base_url)
    telegram_send.send(messages=[msg])

def send_error_notification(tag, alarm_price):
    msg = f"{tag} couldn't found in Binance crypto-list."
    #base_url = f"https://api.telegram.org/bot1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU/sendMessage?chat_id=-{chat_id}&text={msg}"
    #requests.get(base_url)
    telegram_send.send(messages=[msg])

def send_notifications_reply(tag, alarm_price):
    msg = f"{tag} succesfully added to Alarm-List ({alarm_price})."
    telegram_send.send(messages=[msg])
    #base_url = f"https://api.telegram.org/bot1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU/sendMessage?chat_id=-{chat_id}&text={msg}"
    #requests.get(base_url)

def crypto_list():
    client = Client(api_key=api_key, api_secret=api_secret)
    list_crypto = client.get_all_coins_info()
    all_crypto_list = []
    for item in list_crypto:
        all_crypto_list.append(item["coin"])
    return all_crypto_list

def save_alarm_list():
    path_name = "F:/DATABASE/DATA.xlsx"
    if os.path.exists(path_name):
        writer = pd.ExcelWriter(path_name, engine='openpyxl')#pylint: disable=abstract-class-instantiated
        df = pd.DataFrame(alarm_list, columns=["coin_name", "alarm_price", "alarm_type"])
        workbook = writer.book
        try :
            workbook.remove(workbook["Sheet1"])
        except:
            print("Worksheet does not exist ")
        finally:
            df.to_excel(writer, sheet_name="Sheet1",index=False)
            writer.save()
        print("Data has been saved. os path exists")
    else:
        print("An error occured. Path can not be found.")

def load_alarm_list():
    alarm_list_temp = []
    path_name = "F:/DATABASE/DATA.xlsx"
    if os.path.exists(path_name):
        df = pd.read_excel(path_name, index_col=None)
        df["alarm_price"]= df["alarm_price"].astype(str)
        alarm_list_readed = df.values.tolist()
        alarm_list_temp = alarm_list_readed
        print(f"Loaded {alarm_list_temp}")
    else:
        writer = pd.ExcelWriter(path=path_name, engine='xlsxwriter')#pylint: disable=abstract-class-instantiated
        df = pd.DataFrame(alarm_list_temp, columns=["coin_name", "alarm_price", "alarm_type"])
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()
        print("Data can not be loaded. Path doesn't exist. New path has been created.")
    return alarm_list_temp

def load_watch_list():
    watch_list_temp = []
    path_name = "F:/DATABASE/DATA_watchlist.txt"
    if os.path.exists(path_name):
        with open(path_name, "rb") as fp:
            watch_list_temp = pickle.load(fp)
            print(f"Data has been loaded from {path_name}")
    else:
        print(f"{path_name} doesn't exist. A new path has been created.")
    return watch_list_temp

def save_watch_list():
    path_name = "F:/DATABASE/DATA_watchlist.txt"
    with open(path_name, "wb") as fp:
        pickle.dump(watch_list, fp)
    print(f"Data of watch-list has been saved to {path_name}")

def add_tracker(update, context):
    crypto_symbol, time_interval = context.args[0], context.args[1]
    error_handler_var = False
    msg = ""
    if time_interval == "15":
        time_interval_object = Interval.INTERVAL_15_MINUTES
        crpyto_var = crypto_watch(symbol=crypto_symbol, time_interval=time_interval_object)
        watch_list.append(crpyto_var)
        msg = f"{crypto_symbol} succesfully added to watch-list in {time_interval} interval."
        send_analysis_notification(msg, error_handler_var)
    elif time_interval == "1H":
        time_interval_object = Interval.INTERVAL_1_HOUR
        crpyto_var = crypto_watch(symbol=crypto_symbol, time_interval=time_interval_object)
        watch_list.append(crpyto_var)
        msg = f"{crypto_symbol} succesfully added to watch-list in {time_interval} interval."
        send_analysis_notification(msg, error_handler_var)
    elif time_interval == "4H":
        time_interval_object = Interval.INTERVAL_4_HOURS
        crpyto_var = crypto_watch(symbol=crypto_symbol, time_interval=time_interval_object)
        watch_list.append(crpyto_var)
        msg = f"{crypto_symbol} succesfully added to watch-list in {time_interval} interval."
        send_analysis_notification(msg, error_handler_var)
    elif time_interval == "1D":
        time_interval_object = Interval.INTERVAL_1_DAY
        crpyto_var = crypto_watch(symbol=crypto_symbol, time_interval=time_interval_object)
        watch_list.append(crpyto_var)
        msg = f"{crypto_symbol} succesfully added to watch-list in {time_interval} interval."
        send_analysis_notification(msg, error_handler_var)
    else:
        error_handler_var = True
        send_error_notification(msg, error_handler_var)

def review_trackers(update, context):
    for item in watch_list:
        if isinstance(item, crypto_watch):
            msg = f"{item.symbol} in {item.time_interval}"
            telegram_send.send(messages=[msg])
            base_url = f"https://api.telegram.org/bot1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU/sendMessage?chat_id=-{chat_id}&text={msg}"
            requests.get(base_url)

def send_analysis_notification(msg, error_handler_var):
    if error_handler_var:
        msg_error = "An error occured. Please try a validated symbol or interval.(15, 1H, 4H, 1D)"
        telegram_send.send(messages=[msg_error])
        base_url = f"https://api.telegram.org/bot1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU/sendMessage?chat_id=-{chat_id}&text={msg_error}"
        requests.get(base_url)
    elif not error_handler_var:
        base_url = f"https://api.telegram.org/bot1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU/sendMessage?chat_id=-{chat_id}&text={msg}"
        requests.get(base_url)
        telegram_send.send(messages=[msg])

def start(update, context):
    update.message.reply_text("MERHABA, BEN ALLAH 1.0 \n COK ZEKI OLDUM AMK \n --- ONLY WORKS ON BINANCE LISTED CRYPTO CURRENCIES.")

def main():
    #token : 1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU
    Thread(target=check_alarms, daemon=True).start()
    Thread(target=check_watch_list, daemon=True).start()
    logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',level=logging.INFO)
    updater = Updater (token="1891860474:AAFDuHEwzqwuoCNOm-rm7bhS69a9Fab2TpU", use_context=True)
    dispatcher = updater.dispatcher
    start_handler = CommandHandler("start", start)
    add_alarm_handler = CommandHandler("add_alarm", add_new_alarm)
    remove_alarm_handler = CommandHandler("remove_alarm", remove_alarm)
    alarm_reviewer_handle = CommandHandler("review_alarms", review_alarms)
    analysis_handler = CommandHandler("add_tracker", add_tracker)
    analysis_remover_handler = CommandHandler("remove_tracker", remove_from_watch_list)
    analysis_reviewer_handler = CommandHandler("review_trackers", review_trackers)
    atexit.register(save_alarm_list)
    atexit.register(save_watch_list)
    dispatcher.add_handler(analysis_reviewer_handler)
    dispatcher.add_handler(analysis_remover_handler)
    dispatcher.add_handler(alarm_reviewer_handle)
    dispatcher.add_handler(remove_alarm_handler)
    dispatcher.add_handler(analysis_handler)
    dispatcher.add_handler(start_handler)
    dispatcher.add_handler(add_alarm_handler)
    updater.start_polling()
    updater.idle()
    

if __name__ == "__main__":
    alarm_list = load_alarm_list()
    watch_list = load_watch_list()
    print(f"LOADED {watch_list}")
    main()

#add_alarm - Adds a alarm to Alarm-List
#review_alarms - Displays alarms in Alarm-List
#start - Starts bot
#remove_alarm - Removes alarm from Alarm-List
#add_tracker - Adds given crypto to Watch List in order to track and gives buy sell signals
#get_analysis - Gives signals about given time interval and crypto symbol
#remove_tracker - Removes tracker from Signal-List
#review_trackers - Review trackers

#chat = -{chat_id}